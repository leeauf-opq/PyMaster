#  _____      _ _   
# |_   _|    (_) |  
#   | | _ __  _| |_ 
#   | || '_ \| | __|
#  _| || | | | | |_ 
#  \___/_| |_|_|\__|
                                    
import pygame
from pygame.locals import *
import random
import pandas as pd
import time
import math
import openpyxl as op
import copy


pygame.init()
pygame.display.set_caption("PyMaster")
width = 1280
height = 720
win = pygame.display.set_mode((width,height))
file = r"card.xlsx"
data = pd.read_excel(file)
deck_ex = pd.read_excel(r"deck.xlsx")
book = op.load_workbook(filename = 'deck.xlsx')
data_deck = book.active

musique = pygame.mixer.music.load("musique.wav")
pygame.mixer.music.set_volume(0.5)
pygame.mixer.music.play(-1)
clock = pygame.time.Clock()
font = pygame.font.Font("upheavtt.ttf", 35)
petite_font = pygame.font.Font("upheavtt.ttf", 20)
big_font = pygame.font.Font("upheavtt.ttf", 76)

def ReloadBDD():
    global deck_ex, book, data_deck
    deck_ex = pd.read_excel(r"deck.xlsx")
    book = op.load_workbook(filename = 'deck.xlsx')
    data_deck = book.active


#  _   _            _       _     _           
# | | | |          (_)     | |   | |          
# | | | | __ _ _ __ _  __ _| |__ | | ___  ___ 
# | | | |/ _` | '__| |/ _` | '_ \| |/ _ \/ __|
# \ \_/ / (_| | |  | | (_| | |_) | |  __/\__ \
#  \___/ \__,_|_|  |_|\__,_|_.__/|_|\___||___/
                                            
x_0 = (width/2)-8
x_1 = (width/2)-8
tour = 1
changed = False
code = ""

# ______            _                                   _ 
# | ___ \          | |                                 | |
# | |_/ / __ _  ___| | ____ _ _ __ ___  _   _ _ __   __| |
# | ___ \/ _` |/ __| |/ / _` | '__/ _ \| | | | '_ \ / _` |
# | |_/ / (_| | (__|   < (_| | | | (_) | |_| | | | | (_| |
# \____/ \__,_|\___|_|\_\__, |_|  \___/ \__,_|_| |_|\__,_|
#                        __/ |                            
#                       |___/

menu_back = pygame.image.load('image/fond.jpg').convert_alpha()
choose_back = pygame.image.load('image/fond.jpg').convert_alpha()
option_back = pygame.image.load('image/fond.jpg').convert_alpha()
play_back = pygame.image.load('image/fond.jpg').convert_alpha()
editor_hero_back = pygame.image.load('image/fond.jpg').convert_alpha()
editor_back = pygame.image.load('image/fond.jpg').convert_alpha()
choose_deck_back = pygame.image.load('image/fond.jpg').convert_alpha()
deck_viewer_back = pygame.image.load('image/fond.jpg').convert_alpha()
end_back = pygame.image.load('image/fond.jpg').convert_alpha()
tuto_back = pygame.image.load('image/fond.jpg').convert_alpha()
current_background = menu_back
vol_general = 0,5


# ______       _   _                                _   _   _ _____ 
# | ___ \     | | | |                              | | | | | |_   _|
# | |_/ /_   _| |_| |_ ___  _ __     __ _ _ __   __| | | | | | | |  
# | ___ \ | | | __| __/ _ \| '_ \   / _` | '_ \ / _` | | | | | | |  
# | |_/ / |_| | |_| || (_) | | | | | (_| | | | | (_| | | |_| |_| |_ 
# \____/ \__,_|\__|\__\___/|_| |_|  \__,_|_| |_|\__,_|  \___/ \___/ 
                                                                  

logo = pygame.image.load('image/logo.png').convert_alpha()                                                                   
play_0 = pygame.image.load('image/play_0.jpg').convert_alpha()
play_1 = pygame.image.load('image/play_1.jpg').convert_alpha()
option_0 = pygame.image.load('image/option_0.jpg').convert_alpha()
option_1 = pygame.image.load('image/option_1.jpg').convert_alpha()
editeur_0 = pygame.image.load('image/editor_0.jpg').convert_alpha()
editeur_1 = pygame.image.load('image/editor_1.jpg').convert_alpha()
retour_0 = pygame.image.load('image/retour_0.jpg').convert_alpha()
retour_1 = pygame.image.load('image/retour_1.jpg').convert_alpha()
exit_0 = pygame.image.load('image/exit_0.jpg').convert_alpha()
exit_1 = pygame.image.load('image/exit_1.jpg').convert_alpha()
fleche_retour_0 = pygame.image.load('image/fleche_retour_0.png').convert_alpha()
fleche_retour_1 = pygame.image.load('image/fleche_retour_1.png').convert_alpha()
volume = pygame.image.load('image/volume.png').convert_alpha()
barre = pygame.image.load('image/barre.png').convert_alpha()
secret = pygame.image.load('image/secret.png').convert_alpha()
cadre = pygame.image.load('image/cadre.png').convert_alpha()
end_turn_0 = pygame.image.load('image/end_turn_0.png').convert_alpha()
end_turn_1 = pygame.image.load('image/end_turn_1.png').convert_alpha()
aide_0 = pygame.image.load('image/aide_0.png').convert_alpha()
aide_1 = pygame.image.load('image/aide_1.png').convert_alpha()
tuto = pygame.image.load('image/tuto.png').convert_alpha()
parchemin = pygame.image.load('image/parchemin.png').convert_alpha()

#  _   _                          
# | | | |                         
# | |_| | ___ _ __ ___   ___  ___ 
# |  _  |/ _ \ '__/ _ \ / _ \/ __|
# | | | |  __/ | | (_) |  __/\__ \
# \_| |_/\___|_|  \___/ \___||___/
                                
                                
guerrier_0 = pygame.image.load('image/guerrier_0.png').convert_alpha()
guerrier_1 = pygame.image.load('image/guerrier_1.png').convert_alpha()
guerrier = pygame.image.load('image/guerrier.png').convert_alpha()

pretre_0 = pygame.image.load('image/pretre_0.png').convert_alpha()
pretre_1 = pygame.image.load('image/pretre_1.png').convert_alpha()
pretre = pygame.image.load('image/pretre.png').convert_alpha()

mecanicien_0 = pygame.image.load('image/mecanicien_0.png').convert_alpha()
mecanicien_1 = pygame.image.load('image/mecanicien_1.png').convert_alpha()
mecanicien = pygame.image.load('image/mecanicien.png').convert_alpha()

voleur_0 = pygame.image.load('image/voleur_0.png').convert_alpha()
voleur_1 = pygame.image.load('image/voleur_1.png').convert_alpha()
voleur = pygame.image.load('image/voleur.png').convert_alpha()

mage_0 = pygame.image.load('image/mage_0.png').convert_alpha()
mage_1 = pygame.image.load('image/mage_1.png').convert_alpha()
mage = pygame.image.load('image/mage.png').convert_alpha()

classe_ = ["guerrier", "pretre", "mage", "voleur", "mecanicien"]

class Hero:

    def __init__(self, classe, name):        

        self.classe = classe
        self.name = name
        self.stamina = 1
        self.fatigue = 0
        self.current_stamina = self.stamina
        self.blocked = False
        self.boost_damage = 1
        self.heroic_boost_damage = 0
        self.boost_heal = 1
        self.boost_def = 1
        self.damage_taken = 0
        self.esquive = 0
        self.stamina_increase = 0
        self.hp_healed = 0
        self.power_used = False
        self.secret_neutre1 = False
        self.secret_neutre2 = False
        self.secret_neutre3 = False
        self.secret_mage = False
        self.secret_guerrier = False
        self.secret_voleur = False
        self.secret_mecanicien = False
        self.secret_pretre = False
        self.can_play = False

        if classe == "mage":
            self.hp = 15
            self.armor = 0
            self.class_number = 25
            self.deck = []
            self.hand = []
            self.image = mage
            self.class_power_small = pygame.image.load('sprite/pouvoir_mage.png').convert_alpha()
            self.class_power_big = pygame.image.load('sprite/pouvoir_mage1.png').convert_alpha()
            self.class_power_small_off = pygame.image.load('sprite/pouvoir_mage_off.png').convert_alpha()
            self.class_power_big_off = pygame.image.load('sprite/pouvoir_mage1_off.png').convert_alpha()

        if classe == "guerrier":
            self.hp = 15
            self.armor = 0
            self.class_number = 30
            self.deck = []
            self.hand = []
            self.image = guerrier
            self.class_power_small = pygame.image.load('sprite/pouvoir_guerrier.png').convert_alpha()
            self.class_power_big = pygame.image.load('sprite/pouvoir_guerrier1.png').convert_alpha()
            self.class_power_small_off = pygame.image.load('sprite/pouvoir_guerrier_off.png').convert_alpha()
            self.class_power_big_off = pygame.image.load('sprite/pouvoir_guerrier1_off.png').convert_alpha()

        if classe == "voleur":
            self.hp = 15
            self.armor = 0
            self.class_number = 35
            self.deck = []
            self.hand = []
            self.image = voleur
            self.class_power_small = pygame.image.load('sprite/pouvoir_voleur.png').convert_alpha()
            self.class_power_big = pygame.image.load('sprite/pouvoir_voleur1.png').convert_alpha()
            self.class_power_small_off = pygame.image.load('sprite/pouvoir_voleur_off.png').convert_alpha()
            self.class_power_big_off = pygame.image.load('sprite/pouvoir_voleur1_off.png').convert_alpha()

        if classe == "mecanicien":
            self.hp = 10
            self.armor = 5
            self.class_number = 40
            self.deck = []
            self.hand = []
            self.image = mecanicien
            self.class_power_small = pygame.image.load('sprite/pouvoir_mecanicien.png').convert_alpha()
            self.class_power_big = pygame.image.load('sprite/pouvoir_mecanicien1.png').convert_alpha()
            self.class_power_small_off = pygame.image.load('sprite/pouvoir_mecanicien_off.png').convert_alpha()
            self.class_power_big_off = pygame.image.load('sprite/pouvoir_mecanicien1_off.png').convert_alpha()

        if classe == "pretre":
            self.hp = 500
            self.armor = 0
            self.class_number = 45
            self.deck = []
            self.hand = []
            self.image = pretre
            self.class_power_small = pygame.image.load('sprite/pouvoir_pretre.png').convert_alpha()
            self.class_power_big = pygame.image.load('sprite/pouvoir_pretre1.png').convert_alpha()
            self.class_power_small_off = pygame.image.load('sprite/pouvoir_pretre_off.png').convert_alpha()
            self.class_power_big_off = pygame.image.load('sprite/pouvoir_pretre1_off.png').convert_alpha()

    def create_deck(self, class_number):
        deck = []
        for i in range(25):
            deck.append(random.randint(0,24))
        for i in range(5):
            deck.append(class_number+i)
        return sorted(deck)
    
    def create_hand(self, deck):
        hand = []
        longueur = len(deck)-1
        for i in range(5):
            rand = random.randint(0,longueur)
            hand.append(deck.pop(rand))
            longueur -= 1
        return hand
    
#  _   _ _____   _____           _   _             
# | | | |_   _| |  __ \         | | (_)            
# | | | | | |   | |  \/ ___  ___| |_ _  ___  _ __  
# | | | | | |   | | __ / _ \/ __| __| |/ _ \| '_ \ 
# | |_| |_| |_  | |_\ \  __/\__ \ |_| | (_) | | | |
#  \___/ \___/   \____/\___||___/\__|_|\___/|_| |_|


class TextBox(object):
    def __init__(self,rect,**kwargs):
        self.rect = pygame.Rect(rect)
        self.texte = []
        self.final = None
        self.rendered = None
        self.render_rect = None
        self.render_area = None
        self.process_kwargs(kwargs)
        self.char_ligne = 0
        self.max_char = 40
        self.y_text = 5
        self.line_number = 0
 
    def process_kwargs(self,kwargs):
        defaults = {
                    "color" : pygame.Color("black"),
                    "font_color" : pygame.Color("white"),
                    "font" : pygame.font.Font("upheavtt.ttf", 15),
                    }
        for kwarg in kwargs:
            if kwarg in defaults:
                defaults[kwarg] = kwargs[kwarg]
            else:
                raise KeyError("TextBox accepts no keyword {}.".format(kwarg))
        self.__dict__.update(defaults)   
 
    def update(self): 
        win.fill(self.color,self.rect)
        if len(self.texte) > 14:
            del self.texte[:3]
        for i, ligne in enumerate(self.texte):
            self.rendered = self.font.render(ligne, True, self.font_color)
            self.render_rect = self.rendered.get_rect(x=self.rect.x+2, y=(i*25)+self.y_text)
            win.blit(self.rendered,self.render_rect,self.render_area)                                
                                        
def redrawGameWindow(background):
    win.blit(background, (0,0))
    
    if background == menu_back:
        Menu_Background()
           
    if background == option_back:
        Option_Background()
            
    if background == choose_back:
        Choose_Background()

    if background == play_back:
        Play_Background()

    if background == editor_hero_back:
        Editor_Hero_Background()

    if background == editor_back:
        Editor_Background()

    if background == choose_deck_back:
        Choose_Deck_Background()

    if background == deck_viewer_back:
        Deck_Viewer_Background()

    if background == end_back:
        End_Background()

    if background == tuto_back:
        Tuto_Background()
    
    pygame.display.update()

def Menu_Background():
    win.blit(logo, ((width-350)/2,20))
    win.blit(play_0, ((width-400)/4,300))
    win.blit(editeur_0, ((width-400)*(3/4),300))
    win.blit(option_0, ((width-400)/4,500))
    win.blit(exit_0, ((width-400)*(3/4),500))
    win.blit(aide_0, (1150, 10))
    mouse_x, mouse_y = pygame.mouse.get_pos()

    if 1150 <= mouse_x <= 1250 and 20 <= mouse_y <= 105:
        win.blit(aide_1, (1150, 10))
    
    if 300 <= mouse_y <= 420:
        if 220 <= mouse_x <= 620:
            win.blit(play_1, ((width-400)/4,300))
        if 660 <= mouse_x <= 1060:
            win.blit(editeur_1, ((width-400)*(3/4),300))
            
    if 500 <= mouse_y <= 620:
        if 220 <= mouse_x <= 620:
            win.blit(option_1, ((width-400)/4,500))
        if 660 <= mouse_x <= 1060:
            win.blit(exit_1, ((width-400)*(3/4),500))
            

def Option_Background():
    global x_0, x_1
    win.blit(retour_0, ((width-400)/2,550))
    win.blit(volume, ((width/2)-310,150))
    win.blit(barre, (x_0,150))

    win.blit(pygame.image.load("image/son.png").convert_alpha(), (525, 80))

    button = pygame.mouse.get_pressed()
    mouse_x, mouse_y = pygame.mouse.get_pos() 
    
    if 440 <= mouse_x <= 840:
        if 550 <= mouse_y <= 670:
            win.blit(retour_1, ((width-400)/2,550))
    
    if (width/2)-305 <= mouse_x <= (width/2)+295 and 150 <= mouse_y <= 197:  
        if button[0] != 0:
            pos = pygame.mouse.get_pos()
            x_0 = pos[0]
            pygame.mixer.music.set_volume(((x_0-335)*6)/3600) 
        button = pygame.mouse.get_pressed()

    if (width/2)-305 <= mouse_x <= (width/2)+295 and 350 <= mouse_y <= 397:    
        if button[0] != 0:
            pos = pygame.mouse.get_pos()
            x_1 = pos[0]
            pygame.mixer.music.set_volume(((x_1-335)*6)/3600)
        button = pygame.mouse.get_pressed()

def Choose_Background():
    win.blit(guerrier_0, (110,150))
    win.blit(voleur_0, (310,390))
    win.blit(pretre_0, (510,150))
    win.blit(mage_0, (710,390))
    win.blit(mecanicien_0, (910,150))
    win.blit(fleche_retour_0, (10,650))
    mouse_x, mouse_y = pygame.mouse.get_pos() 
    if 150 <= mouse_y <= 450:
        if 130 <= mouse_x <= 350:
            win.blit(guerrier_1, (110,150))
        if 530 <= mouse_x <= 750:
            win.blit(pretre_1, (510,150))
        if 930 <= mouse_x <= 1150:
            win.blit(mecanicien_1, (910,150))
    if 400 <= mouse_y <= 700:
        if 330 <= mouse_x <= 550:
            win.blit(voleur_1, (310,390))
        if 730 <= mouse_x <= 950:
            win.blit(mage_1, (710,390))
        
    if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
        win.blit(fleche_retour_1, (10,650))

def Editor_Hero_Background():
    win.blit(guerrier_0, (110,150))
    win.blit(voleur_0, (310,390))
    win.blit(pretre_0, (510,150))
    win.blit(mage_0, (710,390))
    win.blit(mecanicien_0, (910,150))
    win.blit(fleche_retour_0, (10,650))
    mouse_x, mouse_y = pygame.mouse.get_pos() 
    if 150 <= mouse_y <= 450:
        if 130 <= mouse_x <= 350:
            win.blit(guerrier_1, (110,150))
        if 530 <= mouse_x <= 750:
            win.blit(pretre_1, (510,150))
        if 930 <= mouse_x <= 1150:
            win.blit(mecanicien_1, (910,150))
    if 400 <= mouse_y <= 700:
        if 330 <= mouse_x <= 550:
            win.blit(voleur_1, (310,390))
        if 730 <= mouse_x <= 950:
            win.blit(mage_1, (710,390))
            
    if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
        win.blit(fleche_retour_1, (10,650))

def Play_Background():
    log.update()
    win.blit(cadre, (-15,-15))
    win.blit(end_turn_0, (1110,450))
    win.blit(player_1.image, (10,440))
    win.blit(player_2.image, (1030, 10))
    if player_1.power_used == False:
        win.blit(player_1.class_power_small, (280,450))
    else:
        win.blit(player_1.class_power_small_off, (280,450))
    if player_1.hp < 10:
        win.blit(font.render(str(player_1.hp),3,(255,255,255)), (35,650))
    else:
        win.blit(font.render(str(player_1.hp),3,(255,255,255)), (30,650))
    if player_1.armor < 10:
        win.blit(font.render(str(player_1.armor),3,(0,0,0)), (210,637))
    else:
        win.blit(font.render(str(player_1.armor),3,(0,0,0)), (205,635))
    if player_1.current_stamina < 10:
        win.blit(font.render(str(player_1.current_stamina),3,(255,255,255)), (127,453))
    else:
        win.blit(font.render(str(player_1.current_stamina),3,(255,255,255)), (117,453))

    if player_2.hp < 10:
        win.blit(font.render(str(player_2.hp),3,(255,255,255)), (1050,220))
    else:
        win.blit(font.render(str(player_2.hp),3,(255,255,255)), (1053,220))
    if player_2.armor < 10:
        win.blit(font.render(str(player_2.armor),3,(0,0,0)), (1229,210))
    else:
        win.blit(font.render(str(player_2.armor),3,(0,0,0)), (1223,210))
    if player_2.current_stamina < 10:
        win.blit(font.render(str(player_2.current_stamina),3,(255,255,255)), (1147,24))
    else:
        win.blit(font.render(str(player_2.current_stamina),3,(255,255,255)), (1137,24))
    if (
        player_1.secret_neutre1 == True
        or player_1.secret_neutre2 == True
        or player_1.secret_neutre3 == True
        or player_1.secret_mage == True
        or player_1.secret_guerrier == True
        or player_1.secret_voleur == True
        or player_1.secret_mecanicien == True
        or player_1.secret_pretre == True
    ):
        win.blit(secret, (115, 650))
    if (
    player_2.secret_neutre1 == True
    or player_2.secret_neutre2 == True
    or player_2.secret_neutre3 == True
    or player_2.secret_mage == True
    or player_2.secret_guerrier == True
    or player_2.secret_voleur == True
    or player_2.secret_mecanicien == True
    or player_2.secret_pretre == True
    ):
        win.blit(secret, (1130, 220))
        
    i = 1
    j = 1
    mouse_x, mouse_y = pygame.mouse.get_pos()
    card_blit = -1
    for i, card in enumerate(player_1.hand):
        win.blit(pygame.image.load(data["petit_sprite"][card]).convert_alpha(), (100*i+280, 600))
        if 600 <= mouse_y <= 800:
            if 100*i+280 <= mouse_x <= (100*i+280)+100:
                card_blit = card
                pos_card = i
    if card_blit >= 0:
        if pos_card == 9:
            win.blit(pygame.image.load(data["grand_sprite"][card_blit]).convert_alpha(), (100*pos_card+280-120, 420))
        else: 
            win.blit(pygame.image.load(data["grand_sprite"][card_blit]).convert_alpha(), (100*pos_card+280-75, 420))
                
    for j, card in enumerate(player_2.hand):
        win.blit(pygame.image.load("image/carte_ennemi.png").convert_alpha(), ((width-330)-60*j, 0))

    if 280 < mouse_x < 370 and 460 < mouse_y < 580:
        if player_1.power_used == False:
            win.blit(player_1.class_power_big, (205,347))
        else:
            win.blit(player_1.class_power_big_off, (205,347))

def Editor_Background():
    global nom_deck, created_deck, current_background, deck_edited, code
    mouse_x, mouse_y = pygame.mouse.get_pos()
    win.blit(pygame.image.load("image/recherche.png").convert_alpha(), (740, 10))
    win.blit(pygame.image.load("image/ok.png").convert_alpha(), (1100, 660))
    win.blit(parchemin, (740, 110))
    text = font.render("Supprimer", True, (255,255,255), (0,0,0)) 
    textRect = text.get_rect()
    textRect.center = (950, 680)
    win.blit(text, textRect)

    if "27327327427427627527627598113" in code:
        win.blit(pygame.image.load(data["petit_sprite"][51]).convert_alpha(), (420,440))
            
    if 1200 < mouse_x < 1260 and 650 < mouse_y < 700:
        win.blit(pygame.image.load("image/fleche_retour_1.png").convert_alpha(), (1200, 650))
    else:
        win.blit(pygame.image.load("image/fleche_retour_0.png").convert_alpha(), (1200, 650))       
    card = 0
    for i in range(4):
        for j in range(7):
            if card < 25:
                win.blit(pygame.image.load(data["petit_sprite"][card]).convert_alpha(), (j*100+20,i*140+20))
                card += 1
    for i in range(5):
        win.blit(pygame.image.load(data["petit_sprite"][hero_editor.class_number+i]).convert_alpha(), (i*100+20,580))
    for i in range(4):
        if i*140+20 <= mouse_y <= i*140+136:
            for j in range(7):
                if j*100+20 <= mouse_x <= j*100+120 and j+(7*i) < 25:
                    if i == 0:
                        if j == 0:
                            win.blit(pygame.image.load(data["grand_sprite"][j+(7*i)]).convert_alpha(), (j*100,i*140))
                        else:
                            win.blit(pygame.image.load(data["grand_sprite"][j+(7*i)]).convert_alpha(), (j*100-60,i*140))
                    else:
                        if j == 0:
                            win.blit(pygame.image.load(data["grand_sprite"][j+(7*i)]).convert_alpha(), (j*100,i*140-80))
                        else:
                            win.blit(pygame.image.load(data["grand_sprite"][j+(7*i)]).convert_alpha(), (j*100-60,i*140-80))
    for i in range(5):
        if 580 <= mouse_y <= 696:
            if i*100+20 <= mouse_x <= i*100+120 and i != 0: 
                win.blit(pygame.image.load(data["grand_sprite"][hero_editor.class_number+i]).convert_alpha(), (i*100-60,390))
            elif i*100+20 <= mouse_x <= i*100+120:
                win.blit(pygame.image.load(data["grand_sprite"][hero_editor.class_number+i]).convert_alpha(), (i*100,390))

    if "27327327427427627527627598113" in code:
        if 420 < mouse_x < 420+136 and 440 < mouse_y < 560:
            win.blit(pygame.image.load(data["grand_sprite"][51]).convert_alpha(), (420-80,340))
    
    for i in created_deck:
        if i < 15:
            if created_deck.count(i) == 1:
                win.blit(petite_font.render(data["nom"][i],1,(0,0,0)), (770,i*30+130))
            else:
                win.blit(petite_font.render(data["nom"][i],1,(255,0,0)), (770,i*30+130))
        elif i < 30 <= 50:
            if created_deck.count(i) == 1:
                win.blit(petite_font.render(data["nom"][i],1,(0,0,0)), (960,(i-15)*30+130))
            else:
                win.blit(petite_font.render(data["nom"][i],1,(255,0,0)), (960,(i-15)*30+130))
        elif i == 51:
            win.blit(petite_font.render(data["nom"][i],1,(random.randint(0,255),random.randint(0,255),random.randint(0,255))), (900,570))
        else:
            if created_deck.count(i) == 1:
                win.blit(petite_font.render(data["nom"][i],1,(0,0,255)), (960,(i-hero_editor.class_number+10)*30+130))
            else:
                win.blit(petite_font.render(data["nom"][i],1,(255,0,255)), (960,(i-hero_editor.class_number+10)*30+130))
              
                
    for event in pygame.event.get():
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_BACKSPACE:
                nom_deck = nom_deck[:-1]
            elif len(nom_deck) <= 12:
                nom_deck += event.unicode 
    

        if event.type == pygame.MOUSEBUTTONUP:
            mouse_x, mouse_y = pygame.mouse.get_pos()
            for i in range(4):
                if i*140+20 <= mouse_y <= i*140+136:
                    for j in range(7):
                        if j*100+20 <= mouse_x <= j*100+120 and j+(7*i) < 25:
                            if len(created_deck) < 30:
                                if created_deck.count(j+(7*i)) < 2:
                                    created_deck.append(j+(7*i))

            for i in range(5):
                if 580 <= mouse_y <= 696:
                    if i*100+20 <= mouse_x <= i*100+120:
                        if len(created_deck) < 30:
                            if created_deck.count(hero_editor.class_number+i) < 2:
                                created_deck.append(hero_editor.class_number+i)

            for i in range(15):
                if 770 <= mouse_x <= 940:
                    if i*30+120 <= mouse_y <= i*30+140:
                        if i in created_deck:
                            created_deck.pop(created_deck.index(i))
                if 960 <= mouse_x <= 1180:
                    if i*30+120 <= mouse_y <= i*30+140 and i < 10:
                        if i+15 in created_deck:
                            created_deck.pop(created_deck.index(i+15))
                    elif i*30+120 <= mouse_y <= i*30+140:
                        if hero_editor.class_number+i-10 in created_deck:
                            created_deck.pop(created_deck.index(hero_editor.class_number+i-10))
                            
            if "273" in code:
                if 420 < mouse_x < 420+136 and 440 < mouse_y < 540 and len(created_deck) < 30 and created_deck.count(51) == 0:
                    created_deck.append(51)
                if 900 < mouse_x < 990 and 575 < mouse_y < 585 and 51 in created_deck:
                    created_deck.pop(created_deck.index(51))
    


            if 1100 <= mouse_x <= 1170 and 670 <= mouse_y <= 710:
                if len(created_deck) == 30:
                    if nom_deck == "":
                        nom_deck = f"{hero_editor.classe} new"
                    created_deck.insert(0,nom_deck)
                    created_deck.insert(0,hero_editor.classe)
                    if deck_edited < 0:
                        data_deck.append(created_deck)
                        book.save('deck.xlsx')
                        ReloadBDD()
                        current_background = menu_back
                    else:
                        for i, cell in enumerate(data_deck[str(deck_edited+2)]):
                            cell.value = created_deck[i]
                        book.save('deck.xlsx')
                        ReloadBDD()
                        current_background = menu_back

                    
            if 1200 <= mouse_x <= 1260 and 650 <= mouse_y <= 700:
                current_background = deck_viewer_back
                code = ""
                    

            if 860 <= mouse_x <= 1030 and 660 <= mouse_y <= 700 and deck_edited > 0:
                data_deck.delete_rows(idx=deck_edited+2)
                book.save('deck.xlsx')
                ReloadBDD()
                current_background = deck_viewer_back

    if len(created_deck) == 30:
        win.blit(font.render("Deck complet !",3,(255,255,255)), (700,620))
    else:
        win.blit(font.render(f"Nombres de cartes : {len(created_deck)}",3,(255,255,255)), (540,620))
        
    if len(created_deck) == 0:
        win.blit(font.render(f"Coût moyen : 0",3,(255,255,255)), (980,620))
    else:
        moyen, somme = 0, 0
        try:
            for i in created_deck:
                    somme += int(data["cout"][i])
            moyen = round(somme/len(created_deck), 1)
            win.blit(font.render(f"Coût moyen : {moyen}",3,(255,255,255)), (980,620))
        except:
            pass
    
    win.blit(font.render(nom_deck,2,(0,0,0)), (850,38))


def Choose_Deck_Background():
    win.blit(fleche_retour_0, (10,650))
    mouse_x, mouse_y = pygame.mouse.get_pos()        
    if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
        win.blit(fleche_retour_1, (10,650))
    if player_1.classe == "guerrier":
        win.blit(guerrier_1, (width/2-125,10))
    if player_1.classe == "pretre":
        win.blit(pretre_1, (width/2-125,10))
    if player_1.classe == "mecanicien":
        win.blit(mecanicien_1, (width/2-125,10))
    if player_1.classe == "voleur":
        win.blit(voleur_1, (width/2-125,10))
    if player_1.classe == "mage":
        win.blit(mage_1, (width/2-125,10))
    place = 1
    text = font.render("Vos decks :", True, (255,255,255), (0,0,0)) 
    textRect = text.get_rect()
    textRect.center = (width/2, height/2)
    win.blit(text, textRect)
    for i in range(data_deck.max_row-1):
        if deck_ex['classe'][i] == player_1.classe:
            text = font.render((deck_ex['nom'][i]), True, (255,255,255), (0,0,0)) 
            textRect = text.get_rect()
            if place <= 3:
                textRect.center = (width*(place/4), 500)
            elif place <= 6:
                textRect.center = (width*((place-3)/4), 600)
            win.blit(text, textRect)
            place += 1


def Deck_Viewer_Background():
    global code
    win.blit(fleche_retour_0, (10,650))
    mouse_x, mouse_y = pygame.mouse.get_pos()        
    if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
        win.blit(fleche_retour_1, (10,650))
    if hero_editor.classe == "guerrier":
        win.blit(guerrier_1, (width/2-125,10))
    if hero_editor.classe == "pretre":
        win.blit(pretre_1, (width/2-125,10))
    if hero_editor.classe == "mecanicien":
        win.blit(mecanicien_1, (width/2-125,10))
    if hero_editor.classe == "voleur":
        win.blit(voleur_1, (width/2-125,10))
    if hero_editor.classe == "mage":
        win.blit(mage_1, (width/2-125,10))
    place = 1
    text = font.render("Editer vos decks :", True, (255,255,255), (0,0,0)) 
    textRect = text.get_rect()
    textRect.center = (width/2, height/2)
    win.blit(text, textRect)
    for i in range(data_deck.max_row-1):
        if deck_ex['classe'][i] == hero_editor.classe:
            text = font.render((deck_ex['nom'][i]), True, (255,255,255), (0,0,0)) 
            textRect = text.get_rect()
            if place <= 3:
                textRect.center = (width*(place/4), 450)
            elif place <= 6:
                textRect.center = (width*((place-3)/4), 550)
            win.blit(text, textRect)
            place += 1
    text = font.render("Créer un nouveau Deck", True, (255,255,255), (0,0,0)) 
    textRect = text.get_rect()
    textRect.center = (width/2, 600)
    win.blit(text, textRect)



def End_Background():
    if player_1.hp <= 0:
        if player_2.hp <= 0:
            text = big_font.render("Egalité !", True, (255,255,255), (0,0,0)) 
        else:
            text = big_font.render("P2 a gagné !", True, (255,255,255), (0,0,0)) 
    else:
        text = big_font.render("P1 a gagné !", True, (255,255,255), (0,0,0)) 
    textRect = text.get_rect()
    textRect.center = (width/2, height/2-100)
    win.blit(text, textRect)
    retour = font.render("Menu Principal", True, (255,255,255), (0,0,0)) 
    retourRect = retour.get_rect()
    retourRect.center = (width/2, height/2)
    win.blit(retour, retourRect)

def Tuto_Background():
    win.blit(tuto, (200,100))
    win.blit(fleche_retour_0, (10,650))    
    mouse_x, mouse_y = pygame.mouse.get_pos()
    if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
        win.blit(fleche_retour_1, (10,650))

                            
#  _____                                            _                 _          
# |  __ \                                          | |               (_)         
# | |  \/ __ _ _ __ ___   ___   _ __ ___   ___  ___| |__   __ _ _ __  _  ___ ___ 
# | | __ / _` | '_ ` _ \ / _ \ | '_ ` _ \ / _ \/ __| '_ \ / _` | '_ \| |/ __/ __|
# | |_\ \ (_| | | | | | |  __/ | | | | | |  __/ (__| | | | (_| | | | | | (__\__ \
#  \____/\__,_|_| |_| |_|\___| |_| |_| |_|\___|\___|_| |_|\__,_|_| |_|_|\___|___/
                                                                               
def game(player_1, player_2):
    global tour, first, second
    player = [player_1, player_2]
    rand = random.randint(0,1)
    first = player[rand]
    second = player[not(rand)]
    second.hand.append(50)
    first.can_play = True
    if first == player_2:
        bot_turn(player_2, player_1)

def play(carte, current_player, other_player):
    can_play = True
    if current_player.current_stamina < int(data["cout"][carte]):
        can_play = False
    if can_play == True:
        if "Secret" in data['effet'][carte]:
            pass
        else:
            log.texte.append(f''' -> {current_player.name} utilise la carte "{data['nom'][carte]}" :''')
        for i in range (data['nmb_effet'][carte]):
            if current_player.blocked == True and ( data[f'effet_{i}'][carte] == "heal" or data[f'effet_{i}'][carte] == "armor"):
                log.texte.append(f"{current_player.name} ne peut pas utiliser")
                log.texte.append(f"de heal ou d'armure pendant")
                log.texte.append(f"ce tour")
                return
        try:
            current_player.hand.pop(current_player.hand.index(carte))
        except:
            return
        for j in range (data['nmb_effet'][carte]):
            effect(data[f'effet_{j}'][carte],data[f'valeur_effet_{j}'][carte], current_player, other_player)
        current_player.current_stamina -= int(data["cout"][carte]) + current_player.stamina_increase

def bot_turn(bot, player):
    for i in bot.hand:
        for j in range (data["nmb_effet"][i]):
            if (data[f"effet_{j}"][i] == "damage" and data[f"valeur_effet_{j}"][i] >= player.hp) or data[f"effet_{j}"][i] == "kill" and bot.current_stamina >= data["cout"][i]:
                play(i, bot, player)
                    
    if bot.hp + bot.armor < 10:
        for i in bot.hand:
            for j in range (data["nmb_effet"][i]):
                if (data[f"effet_{j}"][i] == "heal" or data[f"effet_{j}"][i] == "armor" or data[f"effet_{j}"][i] == "mult_hp") and bot.current_stamina >= data["cout"][i]:
                    play(i, bot, player)
                    
    endurance = bot.current_stamina
    cout_main = [data["cout"][i] for i in bot.hand]
    tampon_main = sorted(cout_main, reverse = True)
    run = True
    count = 0
    num = 0
    solution = []

    while run:
        card_played = []
        card_index = []
        tampon_endurance = endurance
        for cout in tampon_main:
            if tampon_endurance >= cout:
                tampon_endurance -= cout
                if cout_main.index(cout) in card_index:
                    card_index.append([i for i, n in enumerate(cout_main) if n == cout])
                else:
                    card_index.append(cout_main.index(cout))
                card_played.append(tampon_main[tampon_main.index(cout)])
        if tampon_endurance == 0 + math.floor(count/100):
            run = False
        else:
            tampon_main = sorted(cout_main, reverse = True)
            random.shuffle(tampon_main)
            count += 2
        if count > 1000:
            run = False

    played_card = dict(zip(card_played, card_index))
    for i in played_card.items():
        i = i[1]
        if type(i) == list:
            for j in i:
                solution.append(j)
        else:
            solution.append(i)
    bot_tampon = copy.deepcopy(bot.hand)
    for carte in solution:
        play(bot_tampon[carte], bot, player)         
    end_turn(bot, player)


def end_turn(current_player, other_player):
    global tour
    tour += 0.5
    current_player.blocked = False
    current_player.stamina_increase = 0
    current_player.power_used = False
    current_player.can_play = False
    current_player, other_player = other_player, current_player
    current_player.can_play = True
    draw(1, current_player)
    if tour >= 2:
        if current_player.stamina < 10:
            current_player.stamina += 1
    current_player.current_stamina = current_player.stamina
    if current_player == player_2:
        bot_turn(player_2, player_1)

                                                                           
def effect(effet, number, current_player, other_player):
    number = int(number)
    if effet == "draw":
        draw(number, current_player)
    if effet == "damage":
        damage(number, current_player, other_player)
    if effet == "heal":
        heal(number, current_player)
    if effet == "armor":
        armor(number, current_player)
    if effet == "discard":
        discard(number, current_player)
    if effet == "bloque":
        block(other_player)
    if effet == "boost_damage":
        boost_damage(current_player)
    if effet == "boost_heal":
        boost_heal(current_player)
    if effet == "div_hp":
        div_hp(other_player)
    if effet == "mult_hp":
        mult_hp(current_player)
    if effet == "stamina":
        stamina(current_player)
    if effet == "duplicate":
        duplicate(current_player)
    if effet == "refill":
        refill(current_player)
    if effet == "boost_def":
        boost_def(current_player)
    if effet == "card_damage":
        damage(len(current_player.hand),current_player, other_player)
    if effet == "return_damage":
        damage(current_player.damage_taken ,current_player, other_player)
    if effet == "esquive":
        esquive(current_player)
    if effet == "stamina_destroy":
        stamina_destroy(other_player)
    if effet == "stamina_increase":
        stamina_increase(other_player)
    if effet == "defausse_adv":
        discard(len(other_player.hand), other_player)
    if effet == "boost_armor":
        boost_armor(current_player)
    if effet == "armor_damage":
        damage(current_player.armor, current_player, other_player)
    if effet == "heal_damage":
        damage(current_player.hp_healed, current_player, other_player)
    if effet == "hand_heal":
        heal(len(current_player.hand),current_player)
    if effet == "temp_stamina":
        temp_stamina(current_player)
    if effet == "steal":
        steal(current_player, other_player)
    if effet == "secret_neutre1":    
        secret_neutre1(current_player)
    if effet == "secret_neutre2":
        secret_neutre2(current_player)
    if effet == "secret_neutre3":
        secret_neutre3(current_player)
    if effet == "secret_mage":
        secret_mage(current_player)
    if effet == "secret_guerrier":        
        secret_guerrier(current_player)
    if effet == "secret_voleur":
        secret_voleur(current_player)
    if effet == "secret_mecanicien":
        secret_mecanicien(current_player)
    if effet == "secret_pretre":
        secret_pretre(current_player)
    if effet == "kill":
        kill(other_player)

def draw(number, player):
    if len(player.deck) >= number:
        longueur = len(player.deck)-1
        for i in range(number):
            rand = random.randint(0,longueur)
            if len(player.hand) >= 10:
                log.texte.append(f"La carte {data['nom'][player.deck.pop(rand)]}") 
                log.texte.append(f"à été supprimé de la main de") 
                log.texte.append(f"{player.name} car il a déjà 10 cartes") 
            else:
                player.hand.append(player.deck.pop(random.randint(0,longueur)))
                longueur -= 1
    else:
        player.fatigue += 1
        log.texte.append(f"{player.name} n'a plus de cartes,") 
        log.texte.append(f"il subit {player.fatigue} dégats de fatigue") 
        player.hp -= player.fatigue
                          
def damage(number, current_player, other_player):
    can_attack = True
    if random.randint(0,100)/100 <= other_player.esquive:
        log.texte.append(f"{other_player.name} esquive les dégats !") 
        other_player.esquive = 0
    else:
        if  other_player.secret_neutre1 == True:
            log.texte.append("L'attaque déclenche le secret Divide") 
            log.texte.append("réduisant les dégats par deux") 
            boost_def(other_player)
            other_player.secret_neutre1 = False
            
        elif other_player.secret_neutre3 == True:
            log.texte.append("L'attaque déclenche le secret Gotcha") 
            log.texte.append(f"faisant piocher deux cartes à {other_player.name}") 
            draw(2, other_player)
            other_player.secret_neutre3 = False
            
        elif other_player.secret_mage == True:
            log.texte.append("L'attaque déclenche le secret Répit") 
            log.texte.append("bloquant les dégats") 
            other_player.boost_def = 0
            other_player.secret_mage = False
            
        elif other_player.secret_voleur == True:
            log.texte.append("L'attaque déclenche le secret Pickpocket") 
            log.texte.append(f"faisant voler {other_player.name} une carte de {current_player.name}") 
            steal(other_player, current_player)
            other_player.secret_voleur = False
            
        elif other_player.secret_mecanicien == True:
            log.texte.append("L'attaque déclenche le secret Agglomération") 
            log.texte.append(f"donnant 5 armure à {other_player.name}") 
            armor(5, other_player)
            other_player.secret_mecanicien = False
            
        elif other_player.secret_pretre == True:
            log.texte.append("L'attaque déclenche le secret Thank You") 
            log.texte.append("et transforment la moitié des dégats en soin") 
            heal(math.floor(number * current_player.boost_damage * other_player.boost_def / 2), other_player)
            other_player.secret_pretre = False
            can_attack = False

        if can_attack == True:
            degat = int(math.floor((number + current_player.heroic_boost_damage) * current_player.boost_damage * other_player.boost_def))
            
            if other_player.secret_neutre2 == True:
                log.texte.append("L'attaque déclenche le secret Parade") 
                log.texte.append(f"renvoyant la moitié des dégats subis à {current_player.name}") 
                damage(math.floor(degat/2), other_player, current_player)
                other_player.secret_neutre2 = False
                
            elif other_player.secret_guerrier == True:
                log.texte.append("L'attaque déclenche le secret Contre") 
                log.texte.append(f"renvoyant les dégats subis sur {current_player.name}") 
                damage(math.floor(degat), other_player, current_player)
                other_player.secret_guerrier = False
                
            if other_player.boost_def < 1:
                other_player.boost_def = 1
            if current_player.boost_damage > 1:
                current_player.boost_damage = 1
                
            if other_player.armor > 0:
                _damage = degat - other_player.armor
                if _damage > 0:
                    other_player.armor = 0
                    other_player.hp -= _damage
                else :
                    other_player.armor = -_damage
            else:
                other_player.hp -= degat
            other_player.damage_taken += degat
            if other_player.hp > 0:
                log.texte.append(f"{other_player.name} perd {degat} pv ! Il lui reste {other_player.hp} pv")
                log.texte.append(f"et {other_player.armor} armure")     
                    
def heal(number, player):
    heal = number * player.boost_heal
    if player.boost_heal > 1:
        player.boost_heal = 1
    player.hp += heal
    player.hp_healed += heal
    if player.hp > 20:
        player.hp = 20
    log.texte.append(f"{player.name} récupere {heal} hp") 
    log.texte.append(f"Il en a maintenant {player.hp}") 
    
def armor(number, player):
    player.armor += number
    log.texte.append(f"{player.name} gagne {number} point d'armure") 
    log.texte.append(f"Il en a maintenant {player.armor}") 
                                  
def discard(number, player):
    longueur = len(player.hand)-1
    for i in range(number):
        rand = random.randint(0,longueur)
        log.texte.append(f"La carte {data['nom'][player.hand.pop(rand)]}") 
        log.texte.append(f"à été défaussé de la main de {player.name}") 
        longueur -= 1
                      
def block(other_player):
    other_player.blocked = True
    log.texte.append(f"Les soins et gains d'armures seront bloqués") 
    log.texte.append(f"pendant le tour de {other_player.name}") 

def boost_damage(player):
    player.boost_damage = 2
    log.texte.append(f"Les prochains dégâts de {player.name}") 
    log.texte.append(f"seront multipliés par 2") 
                                   
                                                
def boost_heal(player):
    player.boost_heal = 2
    log.texte.append(f"Les prochains soins de {player.name}") 
    log.texte.append(f"seront multipliés par 2") 
    
def div_hp(player):
    player.hp = math.floor(player.hp/2)
    log.texte.append(f"La vie de {player.name} à été divisé par deux,")
    log.texte.append(f"il lui reste {player.hp} hp !")
          
def mult_hp(player):
    if player.hp >= 10:
        player.hp = 20
    else:
        player.hp *= 2
    log.texte.append(f"La vie de {player.name} à été multiplié par deux,")
    log.texte.append(f"il lui reste {player.hp} hp !")
                                     
def stamina(player):
    if player.stamina == 10:
        log.texte.append(f"{player.name} a déja 10 d'endurance,")
        log.texte.append(f"il pioche une carte.")
        draw(1, player)
    else:
        player.stamina += 1
        log.texte.append(f"{player.name} gagne 1 d'endurance, il en a {player.stamina}")
        
def duplicate(player):
    longueur = len(player.hand)-1
    player.hand.append(player.hand[random.randint(0,longueur)])
    log.texte.append(f"La carte {data['nom'][player.hand[random.randint(0,longueur)]]}") 
    log.texte.append(f"à été dupliqué dans la main de {player.name}")                  
                     
def refill(player):
    player.current_stamina += 9  
    log.texte.append(f"L'entiere de l'endurance")                                        
    log.texte.append(f"de {player.name} à été restitué")
                                            
def boost_def(player):
    player.boost_def = 0.5
    log.texte.append(f"Les prochains dégats subis")
    log.texte.append(f"par {player.name} seront divisé par 2")

def esquive(player):
    player.esquive = 1
    log.texte.append(f"{player.name} esquivera les prochains dégats")

def stamina_destroy(player):
    player.stamina -= 1
    log.texte.append(f"{player.name} a perdu un d'endurance,")
    log.texte.append(f"il lui en reste {player.stamina}")
                                                                               
def stamina_increase(player):
    player.stamina_increase += 2
    log.texte.append(f"Le coût en endurance des cartes")
    log.texte.append(f"de {player.name} est augmentée de 2")
                                                             
def boost_armor(player):
    player.armor *= 2
    log.texte.append(f"L'armure de {player.name} à été multiplié")
    log.texte.append(f"par 2, il en a {player.armor}")

def temp_stamina(player):
    player.current_stamina += 1
    log.texte.append(f"Le boost donne à {player.name} 1")
    log.texte.append(f"d'endurance pendant ce tour")

def steal(current_player, other_player):
    rand = random.randint(0, len(other_player.hand)-1)
    card = other_player.hand.pop(rand)
    current_player.hand.append(card)
    log.texte.append(f"{current_player.name} vole la carte")
    log.texte.append(f"{data['nom'][card]} à {other_player.name}")

def kill(player):
    player.hp = 0
    
def secret_neutre1(player):
    if player.secret_neutre1 != True:
        player.secret_neutre1 = True
        log.texte.append(f"{player.name} active un secret")
    else:
        log.texte.append("Le secret était déjà actif")
        player.current_stamina += 3 + player.stamina_increase
    
def secret_neutre2(player):
    if player.secret_neutre2 != True:
        player.secret_neutre2 = True
        log.texte.append(f"{player.name} active un secret")
    else:
        log.texte.append("Le secret était déjà actif")
        player.current_stamina += 3 + player.stamina_increase
    
def secret_neutre3(player):
    if player.secret_neutre3 != True:
        player.secret_neutre3 = True
        log.texte.append(f"{player.name} active un secret")
    else:
        log.texte.append("Le secret était déjà actif")
        player.current_stamina += 3 + player.stamina_increase
    
def secret_mage(player):
    if player.secret_mage != True:
        player.secret_mage = True
        log.texte.append(f"{player.name} active un secret")
    else:
        log.texte.append("Le secret était déjà actif")
        player.current_stamina += 3 + player.stamina_increase
    
def secret_guerrier(player):
    if player.secret_guerrier != True:
        player.secret_guerrier = True
        log.texte.append(f"{player.name} active un secret")
    else:
        log.texte.append("Le secret était déjà actif")
        player.current_stamina += 3 + player.stamina_increase
    
def secret_voleur(player):
    if player.secret_voleur != True:
        player.secret_voleur = True
        log.texte.append(f"{player.name} active un secret")
    else:
        log.texte.append("Le secret était déjà actif")
        player.current_stamina += 3 + player.stamina_increase
    
def secret_mecanicien(player):
    if player.secret_mecanicien != True:
        player.secret_mecanicien = True
        log.texte.append(f"{player.name} active un secret")
    else:
        log.texte.append("Le secret était déjà actif")
        player.current_stamina += 3 + player.stamina_increase
        
def secret_pretre(player):
    if player.secret_pretre != True:
        player.secret_pretre = True
        log.texte.append(f"{player.name} active un secret")
    else:
        log.texte.append("Le secret était déjà actif")
        player.current_stamina += 3 + player.stamina_increase


# ___  ___      _         _                   
# |  \/  |     (_)       | |                  
# | .  . | __ _ _ _ __   | | ___   ___  _ __  
# | |\/| |/ _` | | '_ \  | |/ _ \ / _ \| '_ \ 
# | |  | | (_| | | | | | | | (_) | (_) | |_) |
# \_|  |_/\__,_|_|_| |_| |_|\___/ \___/| .__/ 
#                                      | |    
#                                      |_|    

log = TextBox(rect=(0,0,400,400))

run = True
while run:
    clock.tick(60)
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            run = False
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_ESCAPE:
                run = False
            if current_background == deck_viewer_back:
                code += str(event.key)
             
        if event.type == pygame.MOUSEBUTTONUP:
            mouse_x, mouse_y = pygame.mouse.get_pos()

            if current_background == menu_back and changed == True:
                if 300 <= mouse_y <= 420:
                    if 220 <= mouse_x <= 620:
                        current_background = choose_back
                        changed = False
                    if 660 <= mouse_x <= 1060:
                        current_background = editor_hero_back
                        changed = False
            
                if 500 <= mouse_y <= 620:
                    if 220 <= mouse_x <= 620:
                        current_background = option_back
                        changed = False
                    if 660 <= mouse_x <= 1060:
                        run = False                 

                if 1150 <= mouse_x <= 1250 and 20 <= mouse_y <= 105:
                    current_background = tuto_back


            if current_background == choose_back and changed == True:
                if 150 <= mouse_y <= 450:          
                    if 130 <= mouse_x <= 350:
                        player_1 = Hero("guerrier", "P1")
                        player_2 = Hero(classe_[random.randint(0,len(classe_)-1)], "P2")
                        player_2.deck = player_2.create_deck(player_2.class_number)
                        player_2.hand = player_2.create_hand(player_2.deck)
                        current_background = choose_deck_back
                        changed = False
                        
                    if 530 <= mouse_x <= 750:
                        player_1 = Hero("pretre", "P1")
                        player_2 = Hero(classe_[random.randint(0,len(classe_)-1)], "P2")
                        player_2.deck = player_2.create_deck(player_2.class_number)
                        player_2.hand = player_2.create_hand(player_2.deck)
                        current_background = choose_deck_back
                        changed = False
                        
                    if 930 <= mouse_x <= 1150:
                        player_1 = Hero("mecanicien", "P1")
                        player_2 = Hero(classe_[random.randint(0,len(classe_)-1)], "P2")
                        player_2.deck = player_2.create_deck(player_2.class_number)
                        player_2.hand = player_2.create_hand(player_2.deck)
                        current_background = choose_deck_back
                        changed = False
                        
                if 390 <= mouse_y <= 720:    
                    
                    if 330 <= mouse_x <= 550:
                        player_1 = Hero("voleur", "P1")
                        player_2 = Hero(classe_[random.randint(0,len(classe_)-1)], "P2")
                        player_2.deck = player_2.create_deck(player_2.class_number)
                        player_2.hand = player_2.create_hand(player_2.deck)
                        current_background = choose_deck_back
                        changed = False
                        
                    if 730 <= mouse_x <= 950:
                        player_1 = Hero("mage", "P1")
                        player_2 = Hero(classe_[random.randint(0,len(classe_)-1)], "P2")
                        player_2.deck = player_2.create_deck(player_2.class_number)
                        player_2.hand = player_2.create_hand(player_2.deck)
                        current_background = choose_deck_back
                        changed = False

                if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
                    current_background = menu_back
                    changed = False


            if current_background == option_back and changed == True:
                if (width-400)/2 <= mouse_x <= (width+400)/2 and 550 <= mouse_y <= 670:
                    current_background = menu_back
                    changed = False

            if current_background == play_back and changed == True:
                if player_1.can_play == True:
                    if 280 < mouse_x < 370 and 460 < mouse_y < 580:
                        if player_1.current_stamina >= 2 and player_1.power_used == False:
                            if player_1.classe == "guerrier":
                                player_1.heroic_boost_damage = 1
                            elif player_1.classe == "mage":
                                damage(1, player_1, player_2)
                            elif player_1.classe == "voleur":
                                player_1.esquive = 0.2
                            elif player_1.classe == "pretre":
                                heal(1, player_1)
                            elif player_1.classe == "mecanicien":
                                armor(1, player_1)
                            player_1.current_stamina -= 2
                            player_1.power_used = True

                    if 600 <= mouse_y <= 800:
                        for i in range(1, len(player_1.hand)+1):
                            if 100*i+180 <= mouse_x <= (100*i+180)+100:
                                play(player_1.hand[i-1], player_1, player_2)                    
                    if 450 <= mouse_y <= 600:
                        if 1100 <= mouse_x <= 1260:
                            end_turn(player_1, player_2)

            if current_background == end_back and changed == True:
                if 520 <= mouse_x <= 760 and 340 <= mouse_y <= 380:
                    current_background = menu_back


            if current_background == choose_deck_back and changed == True:
                if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
                    current_background = choose_back
                    changed = False
                number_deck = 0
                for i in range(data_deck.max_row-1):
                    if deck_ex['classe'][i] == player_1.classe:
                        number_deck += 1
                if number_deck > 3:
                    first_line = 4
                    second_line = number_deck - 2
                else:
                    first_line = number_deck+1
                    second_line = 0
                number_deck = 0
                if 485 <= mouse_y <= 520:
                    for i in range(1, first_line):
                        if width*(i/4)-115 <= mouse_x <= width*(i/4)+115:
                            for j in range(data_deck.max_row-1):
                                if deck_ex['classe'][j] == player_1.classe:
                                    number_deck += 1
                                    if number_deck == i:
                                        for k in range(1,31):
                                            player_1.deck.append(deck_ex['carte_'+str(k)][j])
                                        player_1.hand = player_1.create_hand(player_1.deck)
                                        game(player_1, player_2)
                                        current_background = play_back
                                        log.texte = []
                                        changed = False
                if 585 <= mouse_y <= 620:
                    for i in range(1, second_line):
                        if width*(i/4)-115 <= mouse_x <= width*(i/4)+115:
                            for j in range(data_deck.max_row-1):
                                if deck_ex['classe'][j] == player_1.classe:
                                    number_deck += 1
                                    if number_deck == i+3:
                                        deck_edited = j
                                        for k in range(1,31):
                                            player_1.deck.append(deck_ex['carte_'+str(k)][j])
                                        player_1.hand = player_1.create_hand(player_1.deck)
                                        game(player_1, player_2)
                                        current_background = play_back
                                        log.texte = []
                                        changed = False

            if current_background == editor_hero_back and changed == True:
                if 150 <= mouse_y <= 450:          
                    if 130 <= mouse_x <= 350:
                        hero_editor = Hero("guerrier", "")
                        current_background = deck_viewer_back
                        changed = False
                        nom_deck = ""
                        created_deck = []
                                               
                    if 530 <= mouse_x <= 750:
                        hero_editor = Hero("pretre", "")
                        current_background = deck_viewer_back
                        changed = False
                        nom_deck = ""
                        created_deck = []
                                              
                    if 930 <= mouse_x <= 1150:
                        hero_editor = Hero("mecanicien", "")
                        current_background = deck_viewer_back
                        changed = False
                        nom_deck = ""
                        created_deck = []
                        
                elif 400 <= mouse_y <= 700:    
                    
                    if 330 <= mouse_x <= 550:
                        hero_editor = Hero("voleur", "")
                        current_background = deck_viewer_back
                        changed = False
                        nom_deck = ""
                        created_deck = []
                        
                    if 730 <= mouse_x <= 950:
                        hero_editor = Hero("mage", "")
                        current_background = deck_viewer_back
                        changed = False
                        nom_deck = ""
                        created_deck = []



                if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
                    current_background = menu_back
                    changed = False

    

            if current_background == deck_viewer_back and changed == True:

                if width/2-100 <= mouse_x <= width/2+100:
                    if 590 <= mouse_y <= 610:
                        current_background = editor_back
                        changed = False
                created_deck = []
                nom_deck = ""
                deck_edited = -1
                number_deck = 0
                for i in range(data_deck.max_row-1):
                    if deck_ex['classe'][i] == hero_editor.classe:
                        number_deck += 1
                if number_deck > 3:
                    first_line = 4
                    second_line = number_deck - 2
                else:
                    first_line = number_deck+1
                    second_line = 0
                number_deck = 0
                if 435 <= mouse_y <= 465:
                    for i in range(1, first_line):
                        if width*(i/4)-115 <= mouse_x <= width*(i/4)+115:
                            for j in range(data_deck.max_row-1):
                                if deck_ex['classe'][j] == hero_editor.classe:
                                    number_deck += 1
                                    if number_deck == i:
                                        deck_edited = j
                                        for k in range(1,31):
                                            created_deck.append(deck_ex['carte_'+str(k)][j])
                                        nom_deck = deck_ex['nom'][j] 
                                        current_background = editor_back
                                        changed = False
                if 535 <= mouse_y <= 565:
                    for i in range(1, second_line):
                        if width*(i/4)-115 <= mouse_x <= width*(i/4)+115:
                            for j in range(data_deck.max_row-1):
                                if deck_ex['classe'][j] == hero_editor.classe:
                                    number_deck += 1
                                    if number_deck == i+3:
                                        deck_edited = j
                                        for k in range(1,31):
                                            created_deck.append(deck_ex['carte_'+str(k)][j])
                                        nom_deck = deck_ex['nom'][j] 
                                        current_background = editor_back
                                        changed = False

                if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
                    current_background = editor_hero_back
                    changed = False

            if current_background == tuto_back:
                if 10 <= mouse_x <= 70 and 650 <= mouse_y <= 710:
                    current_background = menu_back

        if current_background == play_back:
            if player_1.hp <= 0 or player_2.hp <= 0:
                current_background = end_back
                changed= False
                tour = 1
                
                        
    redrawGameWindow(current_background)
    changed = True

        
pygame.quit()
